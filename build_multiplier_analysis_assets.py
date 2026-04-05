from __future__ import annotations

import json
from copy import copy
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List

import openpyxl
import requests
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import extract_target_role_multiplier_summary as core


OUTPUT_WORKBOOK_NAME = "乘区收益图表.xlsx"
SITE_DIR_NAME = "multiplier_wiki"
SITE_DATA_NAME = "data.js"

THIN_SIDE = Side(style="thin", color="D9DEE7")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="23395D")
SECTION_FILL = PatternFill("solid", fgColor="406E8E")
NOTE_FILL = PatternFill("solid", fgColor="FFF4CC")
LABEL_FILL = PatternFill("solid", fgColor="EEF3F8")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def decimal_to_number(value: Decimal) -> int | float:
    return core.decimal_to_number(value)


def bucket_to_payload(bucket: Dict[str, Decimal]) -> Dict[str, int | float]:
    return {
        column: decimal_to_number(bucket.get(column, Decimal("0")))
        for column in core.ATTACK_SUMMARY_EXPORT_COLUMNS
    }


def merge_bucket(dst: Dict[str, Decimal], src: Dict[str, Decimal]) -> None:
    for key, value in src.items():
        dst[key] = dst.get(key, Decimal("0")) + value


def clone_bucket(src: Dict[str, Decimal]) -> Dict[str, Decimal]:
    return {key: value for key, value in src.items()}


def row_total_bucket(row: core.AttackSummaryRow) -> Dict[str, Decimal]:
    bucket = {column: Decimal("0") for column in core.ATTACK_SUMMARY_EXPORT_COLUMNS}
    merge_bucket(bucket, row.base_bucket)
    merge_bucket(bucket, row.effect_bucket)
    return bucket


def row_has_base_damage(row: core.AttackSummaryRow) -> bool:
    return any(row.base_bucket.get(key, Decimal("0")) != 0 for key in core.BASE_KEYS)


def row_to_payload(
    row: core.AttackSummaryRow,
    row_id: str,
    source_urls: Dict[tuple[str, str, str], str],
) -> Dict[str, Any]:
    return {
        "id": row_id,
        "role": row.role,
        "rowType": row.row_type,
        "sourceCol": row.source_col,
        "sourceName": row.source_name,
        "entity": row.entity,
        "baseText": row.base_text,
        "targetScope": row.target_scope,
        "duration": row.duration,
        "notes": list(dict.fromkeys(row.notes)),
        "baseBucket": bucket_to_payload(row.base_bucket),
        "effectBucket": bucket_to_payload(row.effect_bucket),
        "totalBucket": bucket_to_payload(row_total_bucket(row)),
        "hasBaseDamage": row_has_base_damage(row),
        "sourceUrl": source_urls.get((row.role, row.source_col, row.source_name), ""),
    }


def build_site_payload(src_path: Path) -> Dict[str, Any]:
    entries = core.load_unique_entries(src_path)
    rows = core.build_attack_summary_rows(entries)

    role_meta: Dict[str, Dict[str, str]] = {}
    source_urls: Dict[tuple[str, str, str], str] = {}
    avatar_urls: Dict[str, str] = {}
    for entry in entries:
        role_meta.setdefault(entry.role, {"path": entry.path, "element": entry.element})
        source_urls.setdefault((entry.role, entry.source_col, entry.source_name), entry.source_url)
        if "/avatar/" in entry.source_url and entry.role not in avatar_urls:
            avatar_urls[entry.role] = entry.source_url

    role_base_speed: Dict[str, int] = {}
    for role, avatar_url in avatar_urls.items():
        avatar_id = avatar_url.rstrip("/").split("/")[-1]
        try:
            data = requests.get(
                f"https://sr.yatta.moe/api/v2/cn/avatar/{avatar_id}",
                timeout=20,
                headers={"User-Agent": "Mozilla/5.0"},
            ).json()["data"]
            role_base_speed[role] = int(data["upgrade"][0]["skillBase"]["speedBase"])
        except Exception:
            role_base_speed[role] = 100

    rows_by_role: Dict[str, List[Dict[str, Any]]] = {role: [] for role in core.TARGET_ROLES}
    counters: Dict[str, int] = {}
    for row in rows:
        key = f"{row.role}:{row.source_col}:{row.source_name}:{row.row_type}"
        counters[key] = counters.get(key, 0) + 1
        row_id = f"{key}:{counters[key]}"
        rows_by_role.setdefault(row.role, []).append(row_to_payload(row, row_id, source_urls))

    roles_payload = []
    for role in core.TARGET_ROLES:
        meta = role_meta.get(role, {"path": "", "element": ""})
        role_rows = rows_by_role.get(role, [])
        roles_payload.append(
            {
                "name": role,
                "path": meta["path"],
                "element": meta["element"],
                "isSupport": role in core.SUPPORT_ROLES,
                "baseSpeed": role_base_speed.get(role, 100),
                "rows": role_rows,
            }
        )

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "roles": roles_payload,
        "summaryColumns": list(core.ATTACK_SUMMARY_EXPORT_COLUMNS),
        "defaults": {
            "attackerLevel": 80,
            "enemyLevel": 80,
            "enemyResistance": 20,
            "enemyDamageReduction": 0,
            "enemyBroken": False,
            "actionValueLimit": 350,
            "baseCritRate": 5,
            "baseCritDamage": 50,
        },
        "formulaNotes": {
            "defense": "防御乘区按 1 - DEF / (DEF + 200 + 10*攻击者等级) 折算，其中 DEF 默认为 200 + 10*怪物等级，并叠加减防/无视防御。",
            "resistance": "抗性乘区按 1 - 最终抗性 折算，并将最终抗性限制在 -100% 到 90%。",
            "damageReduction": "减伤乘区按 (1 - 敌方减伤) 折算；若敌人未被击破，额外乘 0.9 的韧性减伤。",
        },
    }


def style_sheet_header(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_WRAP
        cell.border = TABLE_BORDER
    ws.row_dimensions[row_idx].height = 28


def add_chart(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title: str,
    x_title: str,
    y_title: str,
    data_start_row: int,
    data_end_row: int,
    anchor: str,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.height = 9
    chart.width = 18
    chart.style = 10

    x_values = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    primary = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
    chart.add_data(primary, titles_from_data=True)
    chart.set_categories(x_values)
    chart.legend.position = "r"

    secondary = LineChart()
    secondary.y_axis.axId = 200
    secondary.y_axis.title = "每+1%边际收益"
    secondary.y_axis.crosses = "max"
    secondary.height = 9
    secondary.width = 18
    secondary.add_data(Reference(ws, min_col=3, min_row=data_start_row - 1, max_row=data_end_row), titles_from_data=True)
    secondary.set_categories(x_values)

    line_primary = chart.series[0]
    line_primary.graphicalProperties.line.solidFill = "E76F51"
    line_primary.graphicalProperties.line.width = 24000

    line_secondary = secondary.series[0]
    line_secondary.graphicalProperties.line.solidFill = "2A9D8F"
    line_secondary.graphicalProperties.line.width = 22000

    chart += secondary
    ws.add_chart(chart, anchor)


def prepare_curve_sheet(
    wb: Workbook,
    title: str,
    description: str,
    x_header: str,
    values: Iterable[int],
    multiplier_formula: Callable[[int], str],
) -> None:
    ws = wb.create_sheet(title)
    ws.append([x_header, "乘区系数", "每+1%边际收益"])
    style_sheet_header(ws, 1)
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90

    values = list(values)
    first_row = 2
    last_row = first_row + len(values) - 1

    ws["E1"] = "说明"
    ws["E1"].fill = SECTION_FILL
    ws["E1"].font = WHITE_FONT
    ws["E1"].alignment = CENTER
    ws["E1"].border = TABLE_BORDER
    ws.merge_cells("E2:H5")
    ws["E2"] = description
    ws["E2"].fill = NOTE_FILL
    ws["E2"].alignment = LEFT_WRAP
    ws["E2"].border = TABLE_BORDER

    for idx, value in enumerate(values, start=first_row):
        ws.cell(idx, 1).value = value
        ws.cell(idx, 2).value = multiplier_formula(idx)
        if idx < last_row:
            ws.cell(idx, 3).value = f"=B{idx + 1}/B{idx}-1"

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = TABLE_BORDER
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    for col in ["E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    for row_idx in range(2, last_row + 1):
        ws.cell(row_idx, 2).number_format = "0.0000"
        ws.cell(row_idx, 3).number_format = "0.00%"

    add_chart(ws, title, x_header, "乘区系数", first_row, last_row, "E7")


def build_analysis_workbook(output_path: Path) -> None:
    wb = Workbook()
    overview = wb.active
    overview.title = "总览"
    overview.sheet_view.zoomScale = 90
    overview.freeze_panes = "A2"

    overview.append(["工作表", "内容", "默认参数/口径"])
    overview.append(["参数", "统一调整敌我等级、抗性、减伤、击破状态，以及双暴收益基线。", "敌人等级默认 80，基础抗性默认 20%。"])
    overview.append(["属性乘区", "展示攻击/生命/防御这类同区相加乘区的数值-收益曲线。", "每多 1% 属性加成的边际收益。"])
    overview.append(["增伤乘区", "展示增伤区曲线。", "每多 1% 增伤的边际收益。"])
    overview.append(["易伤乘区", "展示易伤区曲线。", "每多 1% 易伤的边际收益。"])
    overview.append(["暴击率乘区", "在给定总暴伤基线下，观察暴击率的收益。", "总暴伤可在参数页调整。"])
    overview.append(["暴击伤害乘区", "在给定总暴击率基线下，观察暴伤的收益。", "总暴击率可在参数页调整。"])
    overview.append(["防御乘区", "按默认怪物等级 80，观察减防/无视防御的收益。", "攻击者等级、怪物等级都可调。"])
    overview.append(["抗性乘区", "按可调的目标基础抗性，观察抗穿/减抗的收益。", "默认基础抗性 20%。"])
    overview.append(["减伤乘区", "按可调的敌方减伤与击破状态，观察减伤乘区。", "未击破时额外乘 0.9 韧性减伤。"])
    style_sheet_header(overview, 1)
    overview.column_dimensions["A"].width = 16
    overview.column_dimensions["B"].width = 48
    overview.column_dimensions["C"].width = 44
    for row in overview.iter_rows(min_row=2, max_row=overview.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = TABLE_BORDER
            cell.alignment = LEFT_WRAP

    params = wb.create_sheet("参数", 1)
    params.sheet_view.zoomScale = 95
    params.append(["参数", "数值", "说明"])
    params.append(["攻击者等级", 80, "用于防御乘区，默认角色 80 级。"])
    params.append(["敌方等级", 80, "用于防御乘区，默认怪物 80 级。"])
    params.append(["目标基础抗性(%)", 20, "用于抗性乘区，可改成 0 / 20 / 40 等常见值。"])
    params.append(["敌方额外减伤(%)", 0, "用于减伤乘区。"])
    params.append(["是否已击破弱点", "否", "填“是”则取消默认 0.9 韧性减伤。"])
    params.append(["暴击率收益基线的总暴伤(%)", 100, "例如总暴伤 100%，则每 1% 暴击率带来的期望收益按此折算。"])
    params.append(["暴伤收益基线的总暴击率(%)", 70, "例如总暴击率 70%，则每 1% 暴伤带来的期望收益按此折算。"])
    params.append(["公式来源", "Prydwen / Honkai: Star Rail Wiki Damage Formula", "仅用于收益曲线工作簿的通用乘区分析。"])
    style_sheet_header(params, 1)
    params.freeze_panes = "A2"
    params.column_dimensions["A"].width = 28
    params.column_dimensions["B"].width = 16
    params.column_dimensions["C"].width = 52
    for row in params.iter_rows(min_row=2, max_row=params.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = TABLE_BORDER
            cell.alignment = LEFT_WRAP if cell.column != 2 else CENTER

    prepare_curve_sheet(
        wb,
        "属性乘区",
        "攻击%、生命%、防御% 这类同区相加的属性增幅，乘区系数统一按 1 + x 计算，边际收益会随当前区间越高而递减。",
        "属性加成(%)",
        range(0, 301),
        lambda row: f"=1+A{row}/100",
    )
    prepare_curve_sheet(
        wb,
        "增伤乘区",
        "增伤区按 1 + 增伤% 计算。图里同时给出当前区间下，再额外多 1% 增伤时的边际收益。",
        "增伤(%)",
        range(0, 301),
        lambda row: f"=1+A{row}/100",
    )
    prepare_curve_sheet(
        wb,
        "易伤乘区",
        "易伤区同样按 1 + 易伤% 计算，可直接对比它与增伤区在不同区间的收益差异。",
        "易伤(%)",
        range(0, 201),
        lambda row: f"=1+A{row}/100",
    )
    prepare_curve_sheet(
        wb,
        "暴击率乘区",
        "期望暴击乘区按 1 + 暴击率 × 总暴伤 计算。该表固定总暴伤基线，由参数页控制。",
        "总暴击率(%)",
        range(0, 101),
        lambda row: f"=1+MIN(A{row},100)/100*参数!$B$7/100",
    )
    prepare_curve_sheet(
        wb,
        "暴击伤害乘区",
        "期望暴击乘区按 1 + 总暴击率 × 暴伤 计算。该表固定总暴击率基线，由参数页控制。",
        "总暴伤(%)",
        range(0, 301),
        lambda row: f"=1+参数!$B$8/100*A{row}/100",
    )
    prepare_curve_sheet(
        wb,
        "防御乘区",
        "按 Star Rail 常用防御公式折算。默认怪物等级 80，减防/无视防御从 0% 到 100% 拉出收益曲线。",
        "减防/无视防御(%)",
        range(0, 101),
        lambda row: (
            f"=1-(((200+10*参数!$B$3)*(1-A{row}/100))/"
            f"(((200+10*参数!$B$3)*(1-A{row}/100))+200+10*参数!$B$2))"
        ),
    )
    prepare_curve_sheet(
        wb,
        "抗性乘区",
        "抗性乘区按 1 - 最终抗性 折算，最终抗性会被限制在 -100% 到 90%。目标基础抗性可在参数页调整。",
        "抗穿/减抗(%)",
        range(0, 121),
        lambda row: f"=1-MAX(-100,MIN(90,参数!$B$4-A{row}))/100",
    )
    prepare_curve_sheet(
        wb,
        "减伤乘区",
        "减伤乘区按 (1 - 敌方减伤) 计算；若敌人尚未击破弱点，则额外乘 0.9 的韧性减伤。",
        "敌方减伤(%)",
        range(0, 91),
        lambda row: f"=(1-A{row}/100)*IF(参数!$B$6=\"是\",1,0.9)",
    )

    if output_path.exists():
        output_path.unlink()
    wb.save(output_path)


def build_site_data(src_path: Path, site_dir: Path) -> None:
    payload = build_site_payload(src_path)
    site_dir.mkdir(parents=True, exist_ok=True)
    data_path = site_dir / SITE_DATA_NAME
    content = "window.MULTIPLIER_WIKI_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    data_path.write_text(content, encoding="utf-8")


def main() -> None:
    cwd = Path.cwd()
    src_path = cwd / "角色乘区拆分.xlsx"
    summary_output = cwd / "目标角色乘区汇总.xlsx"
    workbook_output = cwd / OUTPUT_WORKBOOK_NAME
    site_dir = cwd / SITE_DIR_NAME

    core.build_output(src_path, summary_output)
    build_analysis_workbook(workbook_output)
    build_site_data(src_path, site_dir)

    print(summary_output)
    print(workbook_output)
    print(site_dir / SITE_DATA_NAME)


if __name__ == "__main__":
    main()
