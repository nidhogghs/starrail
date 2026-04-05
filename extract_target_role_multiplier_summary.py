from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import openpyxl
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TARGET_ROLES = [
    "姬子",
    "希儿",
    "景元",
    "丹恒•饮月",
    "卡芙卡",
    "流萤",
    "遐蝶",
    "白厄",
    "花火",
    "阮•梅",
    "布洛妮娅",
    "火花",
    "爻光",
]

SUPPORT_ROLES = {"花火", "阮•梅", "布洛妮娅"}

SOURCE_ORDER = [
    "普攻",
    "战技",
    "终结技",
    "天赋",
    "秘技",
    "忆灵/召唤物",
    "额外能力1",
    "额外能力2",
    "额外能力3",
    "属性行迹汇总",
    "专属光锥",
    "E1",
    "E2",
    "E3",
    "E4",
    "E5",
    "E6",
]

EIDOLON_TO_ALLOWED_SOURCES = {
    "0命": {
        "普攻",
        "战技",
        "终结技",
        "天赋",
        "秘技",
        "忆灵/召唤物",
        "额外能力1",
        "额外能力2",
        "额外能力3",
        "属性行迹汇总",
        "专属光锥",
    },
    "2命": {
        "普攻",
        "战技",
        "终结技",
        "天赋",
        "秘技",
        "忆灵/召唤物",
        "额外能力1",
        "额外能力2",
        "额外能力3",
        "属性行迹汇总",
        "专属光锥",
        "E1",
        "E2",
    },
    "6命": set(SOURCE_ORDER),
}

SUMMARY_COLUMNS = [
    "基础倍率_攻击%",
    "基础倍率_生命%",
    "基础倍率_防御%",
    "基础倍率_欢愉伤害%",
    "基础倍率_原伤害%",
    "基础倍率_真实伤害%",
    "基础倍率_击破特攻系数",
    "基础倍率_欢愉度系数",
    "基础倍率_其他%",
    "攻击%",
    "生命%",
    "防御%",
    "固定攻击",
    "固定生命",
    "固定防御",
    "速度%",
    "速度点",
    "增伤区%",
    "易伤区%",
    "暴击率%",
    "暴击伤害%",
    "防御穿透/减防%",
    "抗穿/减抗%",
    "击破特攻%",
    "击破伤害提高%",
    "超击破伤害提高%",
    "弱点击破效率/无视弱点削韧%",
    "欢愉度%",
    "笑点",
    "增笑%",
    "伤害减免/受击减伤%",
]

NUM_RE = r"(\d+(?:\.\d+)?)"


@dataclass
class Entry:
    role: str
    path: str
    element: str
    source_col: str
    source_name: str
    text: str
    source_url: str
    original_categories: List[str] = field(default_factory=list)
    extracted: Dict[str, Decimal] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)


@dataclass
class AttackSummaryRow:
    role: str
    row_type: str
    source_col: str
    source_name: str
    base_text: str
    tags: set[str]
    entity: str = ""
    target_scope: str = ""
    duration: str = ""
    base_bucket: Dict[str, Decimal] = field(default_factory=dict)
    effect_bucket: Dict[str, Decimal] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)


LIGHT_CONE_OVERRIDES = {
    "希儿": [
        (
            "于夜色中",
            "使装备者的暴击率提高18.0%。",
            "https://sr.yatta.moe/api/v2/cn/equipment/23001",
        ),
        (
            "于夜色中",
            "当装备者在战斗中速度大于100时，每超过10点，普攻和战技造成的伤害提高6.0%，该效果可叠加6层。",
            "https://sr.yatta.moe/api/v2/cn/equipment/23001",
        ),
        (
            "于夜色中",
            "当装备者在战斗中速度大于100时，每超过10点，终结技的暴击伤害提高12.0%，该效果可叠加6层。",
            "https://sr.yatta.moe/api/v2/cn/equipment/23001",
        ),
    ],
    "布洛妮娅": [
        (
            "但战斗还未结束",
            "使装备者的能量恢复效率提高10.0%，并在对我方目标施放终结技时恢复1个战技点。该效果每施放2次终结技可触发1次。",
            "https://sr.yatta.moe/api/v2/cn/equipment/23003",
        ),
        (
            "但战斗还未结束",
            "当装备者施放战技后，使下一个行动的我方其他目标造成的伤害提高30.0%，持续1回合。",
            "https://sr.yatta.moe/api/v2/cn/equipment/23003",
        ),
    ],
}

SUPPORT_ROW_OVERRIDES = {
    ("花火", "战技", "梦游鱼"): {
        "target_scope": "指定我方单体",
        "duration": "1回合（额外能力2延长至目标下一个回合开始）",
        "notes": ["另含行动提前50%。"],
    },
    ("花火", "终结技", "一人千役"): {
        "row_type": "增益",
        "target_scope": "持有【谜诡】的我方目标",
        "duration": "2回合",
        "set_bucket": {"增伤区%": Decimal("30")},
        "notes": ["按天赋最多3层折算为额外增伤 10% × 3 = 30%。"],
    },
    ("花火", "天赋", "叙述性诡计"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "2回合",
        "set_bucket": {"增伤区%": Decimal("18")},
        "notes": ["按最多可叠加3层折算为 6% × 3 = 18%。"],
    },
    ("花火", "额外能力3", "夜想曲"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "在场常驻",
        "set_bucket": {"攻击%": Decimal("15")},
    },
    ("花火", "额外能力3", "夜想曲#量子"): {
        "row_type": "增益",
        "target_scope": "我方量子角色（3量子队时）",
        "duration": "在场常驻",
        "set_bucket": {"攻击%": Decimal("30")},
        "notes": ["按文本最高档 30% 计。"],
    },
    ("花火", "专属光锥", "回到大地的飞行"): {
        "row_type": "增益",
        "target_scope": "技能目标",
        "duration": "3回合",
        "set_bucket": {"增伤区%": Decimal("45")},
        "notes": ["按【圣咏】最多3层折算为 15% × 3 = 45%。"],
    },
    ("阮•梅", "战技", "慢捻抹复挑"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "3回合",
        "set_bucket": {"增伤区%": Decimal("68"), "弱点击破效率/无视弱点削韧%": Decimal("50")},
        "notes": ["已并入额外能力3的上限增幅：32% + 36% = 68%。"],
    },
    ("阮•梅", "终结技", "摇花缎水，沾衣不摘"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "2回合",
        "set_bucket": {"抗穿/减抗%": Decimal("25")},
    },
    ("阮•梅", "天赋", "分型的螺旋"): {
        "row_type": "增益",
        "target_scope": "除自身外队友",
        "duration": "在场常驻",
        "set_bucket": {"速度%": Decimal("10")},
    },
    ("阮•梅", "额外能力1", "物体呼吸中"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "在场常驻",
        "set_bucket": {"击破特攻%": Decimal("20")},
    },
    ("阮•梅", "额外能力3", "落烛照水燃"): {
        "skip": True,
    },
    ("阮•梅", "专属光锥", "镜中故我"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "3回合",
        "set_bucket": {"增伤区%": Decimal("24")},
    },
    ("布洛妮娅", "战技", "作战再部署"): {
        "row_type": "增益",
        "target_scope": "指定我方单体",
        "duration": "1回合",
        "set_bucket": {"增伤区%": Decimal("66")},
        "notes": ["另含解除1个负面效果与立即行动。"],
    },
    ("布洛妮娅", "终结技", "贝洛伯格进行曲"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "2回合",
        "set_bucket": {"攻击%": Decimal("55"), "暴击伤害%": Decimal("20")},
        "notes": ["另含基于布洛妮娅暴击伤害的加成：16% × 布洛妮娅暴伤。"],
    },
    ("布洛妮娅", "秘技", "在旗帜下"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "2回合",
        "set_bucket": {"攻击%": Decimal("15")},
    },
    ("布洛妮娅", "额外能力2", "阵地"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "2回合",
        "set_bucket": {"防御%": Decimal("20")},
    },
    ("布洛妮娅", "额外能力3", "军势"): {
        "row_type": "增益",
        "target_scope": "我方全体",
        "duration": "在场常驻",
        "set_bucket": {"增伤区%": Decimal("10")},
    },
    ("布洛妮娅", "专属光锥", "但战斗还未结束"): {
        "row_type": "增益",
        "target_scope": "下一个行动的我方其他目标",
        "duration": "1回合",
        "set_bucket": {"增伤区%": Decimal("30")},
    },
}


def d(value: str | int | float | Decimal) -> Decimal:
    return Decimal(str(value))


def new_bucket() -> Dict[str, Decimal]:
    return {column: Decimal("0") for column in SUMMARY_COLUMNS}


def add_value(bucket: Dict[str, Decimal], key: str, value: Decimal | int | float) -> None:
    bucket[key] += d(value)


def add_match_values(
    bucket: Dict[str, Decimal],
    key: str,
    text: str,
    pattern: str,
    multiplier: Decimal | int | float = 1,
    group: int = 1,
    flags: int = 0,
) -> None:
    factor = d(multiplier)
    for match in re.finditer(pattern, text, flags):
        add_value(bucket, key, d(match.group(group)) * factor)


def strip_percent_numbers(text: str) -> str:
    return re.sub(r"\d+(?:\.\d+)?%", "", text)


def normalize_speed_trace_anomaly(text: str, bucket: Dict[str, Decimal], notes: List[str]) -> str:
    match = re.fullmatch(rf"速度\+(\d{{2,}})\.0%", text.strip())
    if not match:
        return text

    raw_value = d(match.group(1))
    fixed_speed = raw_value / Decimal("100")
    add_value(bucket, "速度点", fixed_speed)
    notes.append(f"检测到明显异常的 `{text}`，按常规行迹值修正为 `速度+{decimal_to_number(fixed_speed)}`。")
    return ""


def surrounding_clause(text: str, start: int, end: int) -> str:
    left = max(text.rfind("\n", 0, start), text.rfind("。", 0, start), text.rfind("；", 0, start))
    right_candidates = [idx for idx in (text.find("\n", end), text.find("。", end), text.find("；", end)) if idx != -1]
    right = min(right_candidates) if right_candidates else len(text)
    return text[left + 1 : right]


def sum_repeated_hits(text: str, bucket: Dict[str, Decimal]) -> None:
    for match in re.finditer(rf"额外造成(\d+)次伤害[^\n。；]*?每次[^。\n；]*?{NUM_RE}%.*?伤害", text):
        count = d(match.group(1))
        pct = d(match.group(2))
        add_value(bucket, "基础倍率_其他%", count * pct)

    for match in re.finditer(rf"造成(\d+)次伤害[^\n。；]*?每次[^。\n；]*?{NUM_RE}%.*?伤害", text):
        full = match.group(0)
        if "额外造成" in full:
            continue
        count = d(match.group(1))
        pct = d(match.group(2))
        add_value(bucket, "基础倍率_其他%", count * pct)

    for match in re.finditer(rf"额外造成(\d+)次[^。\n；]*?各造成[^。\n；]*?{NUM_RE}%.*?伤害", text):
        count = d(match.group(1))
        pct = d(match.group(2))
        add_value(bucket, "基础倍率_其他%", count * pct)


def extract_base_damage(text: str, bucket: Dict[str, Decimal]) -> None:
    add_match_values(bucket, "基础倍率_攻击%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?{NUM_RE}%攻击力")
    add_match_values(bucket, "基础倍率_攻击%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?攻击力{NUM_RE}%")
    add_match_values(bucket, "基础倍率_生命%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?{NUM_RE}%生命上限")
    add_match_values(bucket, "基础倍率_生命%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?(?:生命上限|生命值上限){NUM_RE}%")
    add_match_values(bucket, "基础倍率_防御%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?{NUM_RE}%防御力")
    add_match_values(bucket, "基础倍率_防御%", text, rf"造成[^。\n；]*?等同于[^。\n；]*?防御力{NUM_RE}%")
    add_match_values(bucket, "基础倍率_欢愉伤害%", text, rf"造成{NUM_RE}%的[^。\n；]*?欢愉伤害")
    add_match_values(bucket, "基础倍率_真实伤害%", text, rf"总伤害值{NUM_RE}%的真实伤害")
    add_match_values(bucket, "基础倍率_原伤害%", text, rf"原伤害{NUM_RE}%")
    add_match_values(bucket, "基础倍率_欢愉度系数", text, rf"{NUM_RE}\*欢愉度")
    add_match_values(bucket, "基础倍率_击破特攻系数", text, rf"{NUM_RE}\*击破特攻")

    for match in re.finditer(rf"伤害倍率(?:额外)?提高(?:原倍率的)?{NUM_RE}%", text):
        add_value(bucket, "基础倍率_其他%", d(match.group(1)))

    for match in re.finditer(rf"伤害为原伤害的{NUM_RE}%", text):
        pct = d(match.group(1))
        if pct >= 100:
            add_value(bucket, "增伤区%", pct - 100)
        else:
            add_value(bucket, "基础倍率_原伤害%", pct)

    for match in re.finditer(rf"伤害为原倍率的{NUM_RE}%", text):
        pct = d(match.group(1))
        if pct >= 100:
            add_value(bucket, "增伤区%", pct - 100)
        else:
            add_value(bucket, "基础倍率_原伤害%", pct)

    for match in re.finditer(rf"伤害倍率提高为{NUM_RE}%", text):
        pct = d(match.group(1))
        if pct >= 100:
            add_value(bucket, "增伤区%", pct - 100)
        else:
            add_value(bucket, "基础倍率_其他%", pct)

    sum_repeated_hits(text, bucket)


def extract_stat_bonuses(text: str, bucket: Dict[str, Decimal]) -> None:
    stacked_attack_pattern = rf"攻击力提高{NUM_RE}%[^。\n；]*?(?:最多)?可?叠加(\d+)[层次]"
    for match in re.finditer(stacked_attack_pattern, text):
        add_value(bucket, "攻击%", d(match.group(1)) * d(match.group(2)))
    text = re.sub(stacked_attack_pattern, "", text)

    add_match_values(bucket, "攻击%", text, rf"攻击力(?:提高|\+){NUM_RE}%")
    add_match_values(bucket, "生命%", text, rf"(?:生命上限|生命值上限)(?:提高|\+){NUM_RE}%")
    add_match_values(bucket, "防御%", text, rf"防御力(?:提高|\+){NUM_RE}%")

    text_no_pct = strip_percent_numbers(text)
    add_match_values(bucket, "固定攻击", text_no_pct, rf"攻击力(?:提高|\+){NUM_RE}")
    add_match_values(bucket, "固定生命", text_no_pct, rf"(?:生命上限|生命值上限)(?:提高|\+){NUM_RE}")
    add_match_values(bucket, "固定防御", text_no_pct, rf"防御力(?:提高|\+){NUM_RE}")


def extract_speed(text: str, bucket: Dict[str, Decimal], notes: List[str]) -> None:
    text = normalize_speed_trace_anomaly(text, bucket, notes)
    add_match_values(bucket, "速度%", text, rf"速度(?:提高|\+){NUM_RE}%")
    text_no_pct = strip_percent_numbers(text)
    add_match_values(bucket, "速度点", text_no_pct, rf"(?:基础)?速度(?:提高|\+){NUM_RE}(?:点)?")


def extract_damage_bonus(text: str, bucket: Dict[str, Decimal]) -> None:
    stacked_vuln_pattern = rf"受到的[^。\n；]*?伤害提高{NUM_RE}%[^。\n；]*?(?:最多)?可?叠加(\d+)[层次]"
    for match in re.finditer(stacked_vuln_pattern, text):
        add_value(bucket, "易伤区%", d(match.group(1)) * d(match.group(2)))
    text = re.sub(stacked_vuln_pattern, "", text)

    add_match_values(bucket, "易伤区%", text, rf"受到的[^。\n；]*?伤害提高{NUM_RE}%")

    per_layer_damage_pattern = rf"每层[^。\n；]*?(?<!暴击)伤害提高{NUM_RE}%[^。\n；]*?(?:最多)?可?叠加(\d+)[层次]"
    for match in re.finditer(per_layer_damage_pattern, text):
        add_value(bucket, "增伤区%", d(match.group(1)) * d(match.group(2)))
    text = re.sub(per_layer_damage_pattern, "", text)

    stacked_damage_pattern = rf"(?<!暴击)伤害提高{NUM_RE}%[^。\n；]*?(?:最多)?可?叠加(\d+)[层次]"
    for match in re.finditer(stacked_damage_pattern, text):
        add_value(bucket, "增伤区%", d(match.group(1)) * d(match.group(2)))
    text = re.sub(stacked_damage_pattern, "", text)

    for match in re.finditer(rf"(?:造成的|属性)?持续伤害提高{NUM_RE}%", text):
        add_value(bucket, "增伤区%", d(match.group(1)))

    for match in re.finditer(rf"(?:[全风火雷冰量子虚数物理]+属性)?(?<!暴击)伤害\+{NUM_RE}%", text):
        clause = match.group(0)
        if "暴击" in clause:
            continue
        add_value(bucket, "增伤区%", d(match.group(1)))

    for match in re.finditer(rf"造成的[^。\n；]*?伤害提高{NUM_RE}%", text):
        clause = match.group(0)
        if "受到的" in clause or "暴击伤害" in clause:
            continue
        add_value(bucket, "增伤区%", d(match.group(1)))

    for match in re.finditer(rf"(?<!受到的)(?<!暴击)(?<!击破)伤害提高{NUM_RE}%", text):
        clause = surrounding_clause(text, match.start(), match.end())
        if "造成的" in clause or "持续伤害" in clause:
            continue
        add_value(bucket, "增伤区%", d(match.group(1)))

    for match in re.finditer(rf"伤害提高的效果额外提高{NUM_RE}%[^。\n；]*?最高(?:不超过|提高){NUM_RE}%", text):
        add_value(bucket, "增伤区%", d(match.group(2)))


def extract_crit(text: str, bucket: Dict[str, Decimal]) -> None:
    add_match_values(bucket, "暴击率%", text, rf"暴击率(?:提高|\+){NUM_RE}%")

    stacked_cd_pattern = rf"暴击伤害提高{NUM_RE}%[^。\n；]*?(?:最多)?可?叠加(\d+)[层次]"
    for match in re.finditer(stacked_cd_pattern, text):
        add_value(bucket, "暴击伤害%", d(match.group(1)) * d(match.group(2)))
    text = re.sub(stacked_cd_pattern, "", text)

    for match in re.finditer(rf"暴击伤害(?:提高|\+){NUM_RE}%", text):
        clause = surrounding_clause(text, match.start(), match.end())
        if "每消耗1个【爆点】" in clause or "每拥有1个笑点" in clause:
            continue
        add_value(bucket, "暴击伤害%", d(match.group(1)))

    for match in re.finditer(rf"每消耗1个【爆点】使自身暴击伤害提高{NUM_RE}%，持续[^。\n；]*?最多叠加(\d+)[层次]", text):
        base = d(match.group(1))
        stacks = d(match.group(2))
        add_value(bucket, "暴击伤害%", base * stacks)

    for match in re.finditer(rf"当前每拥有1个笑点，使[^。\n；]*?暴击伤害提高{NUM_RE}%，最多提高{NUM_RE}%", text):
        add_value(bucket, "暴击伤害%", d(match.group(2)))


def extract_penetration(text: str, bucket: Dict[str, Decimal]) -> None:
    for match in re.finditer(rf"无视[^。\n；]*?{NUM_RE}%的防御力", text):
        clause = surrounding_clause(text, match.start(), match.end())
        if "每消耗1个战技点" in clause and "最多叠加" in clause:
            continue
        add_value(bucket, "防御穿透/减防%", d(match.group(1)))

    add_match_values(bucket, "防御穿透/减防%", text, rf"无视[^。\n；]*?{NUM_RE}%防御力")
    add_match_values(bucket, "防御穿透/减防%", text, rf"防御力降低{NUM_RE}%")

    for match in re.finditer(rf"抗性穿透提高{NUM_RE}%", text):
        clause = surrounding_clause(text, match.start(), match.end())
        if "每拥有1个笑点" in clause:
            continue
        add_value(bucket, "抗穿/减抗%", d(match.group(1)))

    add_match_values(bucket, "抗穿/减抗%", text, rf"全属性抗性降低{NUM_RE}%")
    add_match_values(bucket, "抗穿/减抗%", text, rf"[风火雷冰量子虚数物理]+属性抗性降低{NUM_RE}%")

    for match in re.finditer(rf"每消耗1个战技点使[^。\n；]*?无视敌方目标{NUM_RE}%的防御力[^。\n；]*?最多叠加(\d+)层", text):
        add_value(bucket, "防御穿透/减防%", d(match.group(1)) * d(match.group(2)))


def extract_break(text: str, bucket: Dict[str, Decimal]) -> None:
    add_match_values(bucket, "击破特攻%", text, rf"(?<!\*)击破特攻(?:提高|\+){NUM_RE}%")
    add_match_values(bucket, "击破伤害提高%", text, rf"击破伤害提高{NUM_RE}%")
    add_match_values(bucket, "超击破伤害提高%", text, rf"超击破伤害提高{NUM_RE}%")
    add_match_values(bucket, "弱点击破效率/无视弱点削韧%", text, rf"弱点击破效率提高{NUM_RE}%")
    add_match_values(bucket, "弱点击破效率/无视弱点削韧%", text, rf"无视弱点削韧(?:提高)?{NUM_RE}%")


def extract_elation(text: str, bucket: Dict[str, Decimal], notes: List[str]) -> None:
    for match in re.finditer(rf"欢愉度(?:提高|\+){NUM_RE}%", text):
        clause = surrounding_clause(text, match.start(), match.end())
        if match.start() > 0 and text[match.start() - 1] == "*":
            continue
        if "每超过100点攻击力" in clause or "提高数值等同于" in clause:
            continue
        add_value(bucket, "欢愉度%", d(match.group(1)))

    for match in re.finditer(rf"每超过100点攻击力可使自身欢愉度提高{NUM_RE}%，最多提高{NUM_RE}%", text):
        add_value(bucket, "欢愉度%", d(match.group(2)))

    for match in re.finditer(rf"提高数值等同于[^。\n；]*?欢愉度的{NUM_RE}%", text):
        add_value(bucket, "欢愉度%", d(match.group(1)))

    add_match_values(bucket, "增笑%", text, rf"欢愉伤害增笑{NUM_RE}%")

    for match in re.finditer(rf"(?:获得|额外获得)(\d+)个笑点", text):
        add_value(bucket, "笑点", d(match.group(1)))

    for match in re.finditer(rf"固定计入(\d+)笑点", text):
        add_value(bucket, "笑点", d(match.group(1)))

    for match in re.finditer(rf"当前每拥有1个笑点，使[^。\n；]*?抗性穿透提高{NUM_RE}%，最多提高{NUM_RE}%", text):
        add_value(bucket, "抗穿/减抗%", d(match.group(2)))

    if re.search(r"\d+/\d+/\d+个笑点", text):
        notes.append("存在分档笑点数值，汇总未自动取单一档位。")


def extract_damage_reduction(text: str, bucket: Dict[str, Decimal]) -> None:
    add_match_values(bucket, "伤害减免/受击减伤%", text, rf"受到的伤害降低{NUM_RE}%")
    add_match_values(bucket, "伤害减免/受击减伤%", text, rf"减伤效果达到最大值，最多降低{NUM_RE}%")


def special_case_adjustments(entry: Entry, bucket: Dict[str, Decimal]) -> None:
    text = entry.text

    if entry.role == "希儿" and entry.source_col == "E2" and "最多叠加2层" in text:
        add_value(bucket, "速度%", Decimal("25"))
        entry.notes.append("E2 将战技的 25% 速度提升改为可叠 2 层，已补入额外 25%。")

    if entry.role == "景元" and entry.source_col == "额外能力3" and "暴击率提升10.0%" in text:
        add_value(bucket, "暴击率%", Decimal("10"))

    if entry.role == "爻光" and entry.source_col == "E4" and "欢愉技造成的伤害为原伤害的150.0%" in text:
        add_value(bucket, "增伤区%", Decimal("50"))

    if entry.role == "白厄" and entry.source_col == "额外能力1" and "白厄的暴击伤害提高30.0%" in text:
        add_value(bucket, "暴击伤害%", Decimal("30"))

def extract_entry(entry: Entry) -> None:
    bucket = new_bucket()
    text = entry.text or ""

    extract_base_damage(text, bucket)
    extract_stat_bonuses(text, bucket)
    extract_speed(text, bucket, entry.notes)
    extract_damage_bonus(text, bucket)
    extract_crit(text, bucket)
    extract_penetration(text, bucket)
    extract_break(text, bucket)
    extract_elation(text, bucket, entry.notes)
    extract_damage_reduction(text, bucket)
    special_case_adjustments(entry, bucket)

    entry.extracted = {k: v for k, v in bucket.items() if v != 0}


BASE_KEYS = {
    "基础倍率_攻击%",
    "基础倍率_生命%",
    "基础倍率_防御%",
    "基础倍率_欢愉伤害%",
    "基础倍率_原伤害%",
    "基础倍率_真实伤害%",
    "基础倍率_击破特攻系数",
    "基础倍率_欢愉度系数",
    "基础倍率_其他%",
}


ZERO_EIDOLON_SOURCES = EIDOLON_TO_ALLOWED_SOURCES["0命"]


def extract_bucket_from_text(text: str, role: str = "", source_col: str = "", source_name: str = "") -> Dict[str, Decimal]:
    entry = Entry(role=role, path="", element="", source_col=source_col, source_name=source_name, text=text, source_url="")
    extract_entry(entry)
    return entry.extracted


def non_base_bucket(bucket: Dict[str, Decimal]) -> Dict[str, Decimal]:
    return {k: v for k, v in bucket.items() if k not in BASE_KEYS and v != 0}


def split_effect_clauses(text: str) -> List[str]:
    clauses: List[str] = []
    for part in re.split(r"[\n]+", text):
        part = part.strip()
        if not part:
            continue
        for sub_part in re.split(r"(?:(?:，|；)同时)", part):
            sub_part = sub_part.strip("，；。 ")
            if sub_part:
                clauses.append(sub_part)
    return clauses


def apply_entry_overrides(entries: List[Entry]) -> List[Entry]:
    filtered = [entry for entry in entries if not (entry.role in LIGHT_CONE_OVERRIDES and entry.source_col == "专属光锥")]
    for role, light_cone_entries in LIGHT_CONE_OVERRIDES.items():
        for source_name, text, source_url in light_cone_entries:
            entry = Entry(
                role=role,
                path="",
                element="",
                source_col="专属光锥",
                source_name=source_name,
                text=text,
                source_url=source_url,
                original_categories=["手动修正为角色原始专武"],
            )
            extract_entry(entry)
            entry.notes.append("已替换原先的候选光锥，改按角色原始专武口径汇总。")
            filtered.append(entry)

    order_map = {value: idx for idx, value in enumerate(SOURCE_ORDER)}
    filtered.sort(key=lambda item: (TARGET_ROLES.index(item.role), order_map.get(item.source_col, 999), item.source_name, item.text))
    return filtered

def load_unique_entries(src_path: Path) -> List[Entry]:
    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb["长表"]
    by_key: Dict[Tuple[str, str, str, str, str], Entry] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        role = row[1]
        if role not in TARGET_ROLES:
            continue

        key = (role, row[4], row[6], row[7], row[8])
        if key not in by_key:
            by_key[key] = Entry(
                role=role,
                path=row[2],
                element=row[3],
                source_col=row[4],
                source_name=row[6],
                text=row[7] or "",
                source_url=row[8] or "",
                original_categories=[],
            )

        category = row[5]
        if category and category not in by_key[key].original_categories:
            by_key[key].original_categories.append(category)

    grouped_entries: Dict[Tuple[str, str, str], List[Entry]] = defaultdict(list)
    for entry in by_key.values():
        grouped_entries[(entry.role, entry.source_col, entry.source_name)].append(entry)

    entries: List[Entry] = []
    for group in grouped_entries.values():
        for entry in group:
            text = entry.text.strip()
            redundant = False
            for other in group:
                if other is entry:
                    continue
                other_text = other.text.strip()
                if text and text != other_text and text in other_text and len(other_text) > len(text):
                    redundant = True
                    break
            if not redundant:
                entries.append(entry)

    order_map = {value: idx for idx, value in enumerate(SOURCE_ORDER)}
    entries.sort(key=lambda item: (TARGET_ROLES.index(item.role), order_map.get(item.source_col, 999), item.source_name, item.text))

    for entry in entries:
        extract_entry(entry)

    return apply_entry_overrides(entries)


def summarize(entries: Iterable[Entry]) -> Dict[Tuple[str, str], Dict[str, Decimal]]:
    summary: Dict[Tuple[str, str], Dict[str, Decimal]] = {}

    for role in TARGET_ROLES:
        role_entries = [entry for entry in entries if entry.role == role]
        for eidolon, allowed_sources in EIDOLON_TO_ALLOWED_SOURCES.items():
            bucket = new_bucket()
            for entry in role_entries:
                if entry.source_col not in allowed_sources:
                    continue
                for key, value in entry.extracted.items():
                    bucket[key] += value
            summary[(role, eidolon)] = bucket

    return summary


ATTACK_SUMMARY_EXPORT_COLUMNS = [
    "基础倍率_攻击%",
    "基础倍率_生命%",
    "基础倍率_防御%",
    "基础倍率_欢愉伤害%",
    "基础倍率_原伤害%",
    "基础倍率_真实伤害%",
    "基础倍率_击破特攻系数",
    "基础倍率_欢愉度系数",
    "基础倍率_其他%",
    "攻击%",
    "生命%",
    "防御%",
    "速度%",
    "速度点",
    "增伤区%",
    "易伤区%",
    "暴击率%",
    "暴击伤害%",
    "防御穿透/减防%",
    "抗穿/减抗%",
    "击破特攻%",
    "击破伤害提高%",
    "超击破伤害提高%",
    "弱点击破效率/无视弱点削韧%",
    "欢愉度%",
    "笑点",
    "增笑%",
    "伤害减免/受击减伤%",
]


def bucket_with_defaults() -> Dict[str, Decimal]:
    return {column: Decimal("0") for column in ATTACK_SUMMARY_EXPORT_COLUMNS}


def merge_bucket(dst: Dict[str, Decimal], src: Dict[str, Decimal]) -> None:
    for key, value in src.items():
        if key in dst:
            dst[key] += value


def entry_has_base_damage(entry: Entry) -> bool:
    return any(key in entry.extracted for key in BASE_KEYS) or bool(re.search(r"(持续伤害|欢愉伤害|附加伤害)", entry.text))


def infer_row_type(entry: Entry) -> str:
    if "持续伤害" in entry.text:
        return "持续伤害"
    if entry.source_col == "忆灵/召唤物":
        return "召唤物"
    if entry.source_col == "天赋" and ("追加攻击" in entry.text or "欢愉伤害" in entry.text or "【神君】" in entry.text):
        return "追加攻击"
    if entry.source_col == "天赋" and "附加伤害" in entry.text:
        return "附加伤害"
    return entry.source_col


def infer_tags(row_type: str, source_name: str, text: str, role: str) -> set[str]:
    tags = {row_type}
    if row_type == "召唤物":
        tags.add("召唤物")
        tags.add("死龙")
    if row_type == "追加攻击":
        tags.add("追加攻击")
    if row_type == "持续伤害":
        tags.add("持续伤害")
    if "欢愉伤害" in text:
        tags.add("欢愉伤害")
    if "神君" in text or "神君" in source_name:
        tags.add("神君")
    return tags


def infer_target_scope(text: str) -> str:
    target_patterns = [
        ("下一个行动的我方其他目标", "下一个行动的我方其他目标"),
        ("指定我方单体", "指定我方单体"),
        ("我方单体", "我方单体"),
        ("持有【谜诡】的我方目标", "持有【谜诡】的我方目标"),
        ("我方量子属性的角色", "我方量子角色"),
        ("除自身以外的队友", "除自身外队友"),
        ("技能目标", "技能目标"),
        ("我方全体", "我方全体"),
    ]
    for pattern, label in target_patterns:
        if pattern in text:
            return label
    return ""


def infer_duration(text: str) -> str:
    match = re.search(r"持续(\d+)回合", text)
    if match:
        return f"{match.group(1)}回合"
    if "在场时" in text or "在场" in text:
        return "在场常驻"
    return ""


def support_override_for(entry: Entry) -> Dict[str, object] | None:
    key = support_override_key(entry)
    if key in SUPPORT_ROW_OVERRIDES:
        return SUPPORT_ROW_OVERRIDES[key]
    return None


def support_override_key(entry: Entry) -> tuple[str, str, str]:
    if entry.role == "花火" and entry.source_col == "额外能力3" and "量子属性的角色" in entry.text:
        return ("花火", "额外能力3", "夜想曲#量子")
    return (entry.role, entry.source_col, entry.source_name)


def entry_affects_allies(entry: Entry) -> bool:
    ally_tokens = ["我方", "队友", "技能目标", "指定我方", "持有【谜诡】的我方目标"]
    return any(token in entry.text for token in ally_tokens)


def build_base_rows(entries: Iterable[Entry]) -> List[AttackSummaryRow]:
    rows: List[AttackSummaryRow] = []
    for role in TARGET_ROLES:
        if role in SUPPORT_ROLES:
            continue
        role_entries = [e for e in entries if e.role == role and e.source_col in ZERO_EIDOLON_SOURCES]
        seen: set[tuple[str, str, str]] = set()
        for entry in role_entries:
            if entry.source_col not in {"普攻", "战技", "终结技", "天赋", "秘技", "忆灵/召唤物", "额外能力1", "额外能力2", "额外能力3"}:
                continue
            if not entry_has_base_damage(entry):
                continue
            row_type = infer_row_type(entry)
            key = (role, row_type, entry.source_name)
            if key in seen:
                continue
            seen.add(key)
            base_bucket = bucket_with_defaults()
            merge_bucket(base_bucket, {k: v for k, v in entry.extracted.items() if k in ATTACK_SUMMARY_EXPORT_COLUMNS and k in BASE_KEYS})
            effect_bucket = bucket_with_defaults()
            merge_bucket(effect_bucket, non_base_bucket(entry.extracted))
            rows.append(
                AttackSummaryRow(
                    role=role,
                    row_type=row_type,
                    source_col=entry.source_col,
                    source_name=entry.source_name,
                    base_text=entry.text,
                    tags=infer_tags(row_type, entry.source_name, entry.text, role),
                    entity="死龙" if row_type == "召唤物" else "",
                    target_scope="",
                    duration="",
                    base_bucket=base_bucket,
                    effect_bucket=effect_bucket,
                )
            )
    return rows


def build_support_rows(entries: Iterable[Entry]) -> List[AttackSummaryRow]:
    rows: List[AttackSummaryRow] = []
    support_entries = [entry for entry in entries if entry.role in SUPPORT_ROLES and entry.source_col in ZERO_EIDOLON_SOURCES]
    grouped_entries: Dict[tuple[str, str, str], List[Entry]] = defaultdict(list)
    for entry in support_entries:
        key = support_override_key(entry)
        grouped_entries[key].append(entry)

    for key, group in grouped_entries.items():
        sample = group[0]
        combined_text = "\n".join(entry.text for entry in group if entry.text)
        combined_entry = Entry(
            role=sample.role,
            path=sample.path,
            element=sample.element,
            source_col=sample.source_col,
            source_name=sample.source_name,
            text=combined_text,
            source_url=sample.source_url,
        )
        extract_entry(combined_entry)

        override = SUPPORT_ROW_OVERRIDES.get(key)
        if override and override.get("skip"):
            continue
        include_damage_row = sample.source_col == "普攻" and entry_has_base_damage(combined_entry)
        include_support_row = override is not None or (entry_affects_allies(combined_entry) and bool(non_base_bucket(combined_entry.extracted)))
        if not include_damage_row and not include_support_row:
            continue

        row_type = infer_row_type(combined_entry) if include_damage_row else str((override or {}).get("row_type", "增益"))
        base_bucket = bucket_with_defaults()
        effect_bucket = bucket_with_defaults()
        if include_damage_row:
            merge_bucket(base_bucket, {k: v for k, v in combined_entry.extracted.items() if k in ATTACK_SUMMARY_EXPORT_COLUMNS and k in BASE_KEYS})
            merge_bucket(effect_bucket, non_base_bucket(combined_entry.extracted))
        else:
            merge_bucket(effect_bucket, non_base_bucket(combined_entry.extracted))

        notes = []
        for entry in group:
            notes.extend(entry.notes)
        notes.extend(combined_entry.notes)
        target_scope = infer_target_scope(combined_text)
        duration = infer_duration(combined_text)
        if override:
            if "set_bucket" in override:
                effect_bucket = bucket_with_defaults()
                for bucket_key, value in dict(override["set_bucket"]).items():
                    effect_bucket[bucket_key] = value
            target_scope = str(override.get("target_scope", target_scope))
            duration = str(override.get("duration", duration))
            notes.extend(str(note) for note in override.get("notes", []))

        rows.append(
            AttackSummaryRow(
                role=sample.role,
                row_type=row_type,
                source_col=sample.source_col,
                source_name=sample.source_name,
                base_text=combined_text,
                tags=infer_tags(row_type, sample.source_name, combined_text, sample.role),
                target_scope=target_scope,
                duration=duration,
                base_bucket=base_bucket,
                effect_bucket=effect_bucket,
                notes=list(dict.fromkeys(note for note in notes if note)),
            )
        )
    return rows


def fetch_castorice_servant_rows() -> List[AttackSummaryRow]:
    url = "https://sr.yatta.moe/api/v2/cn/avatar/1407"
    data = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"}).json()["data"]
    servant = data["traces"]["servantSkills"]

    slash = servant["skills"]["1407301"]["skillList"]["1140701"]
    breath = servant["skills"]["1407301"]["skillList"]["1140702"]
    wing = servant["talents"]["1407302"]["skillList"]["1140706"]

    slash_rate = Decimal(str(slash["params"]["1"][5] * 100))
    breath_rate_1 = Decimal(str(breath["params"]["2"][5] * 100))
    breath_rate_2 = Decimal(str(breath["params"]["3"][5] * 100))
    breath_rate_3 = Decimal(str(breath["params"]["4"][5] * 100))
    wing_rate = Decimal(str(wing["params"]["1"][5] * 100))
    wing_hits = int(wing["params"]["2"][5])

    manual_rows = [
        ("召唤物", "擘裂冥茫的爪痕", f"对敌方全体造成等同于遐蝶{slash_rate}%生命上限的量子属性伤害。"),
        ("召唤物", "燎尽黯泽的焰息", f"对敌方全体造成等同于遐蝶{breath_rate_1}% / {breath_rate_2}% / {breath_rate_3}%生命上限的量子属性伤害，重复施放时倍率递增，最高按 {breath_rate_3}% 计。"),
        ("召唤物", "灼掠幽墟的晦翼", f"造成{wing_hits}次伤害，每次对随机敌方单体造成等同于遐蝶{wing_rate}%生命上限的量子属性伤害。"),
    ]

    rows: List[AttackSummaryRow] = []
    for row_type, source_name, text in manual_rows:
        bucket = bucket_with_defaults()
        merge_bucket(bucket, extract_bucket_from_text(text, role="遐蝶", source_col="忆灵/召唤物", source_name=source_name))
        rows.append(
            AttackSummaryRow(
                role="遐蝶",
                row_type=row_type,
                source_col="忆灵/召唤物",
                source_name=source_name,
                base_text=text,
                tags={"召唤物", "死龙", row_type},
                entity="死龙",
                base_bucket={k: bucket.get(k, Decimal("0")) for k in ATTACK_SUMMARY_EXPORT_COLUMNS},
                effect_bucket=bucket_with_defaults(),
                notes=["死龙倍率来自 Yatta 原始 servantSkills，按无额外命座的基准技能等级取值。"],
            )
        )
    return rows


def clause_has_effect(clause: str) -> bool:
    bucket = non_base_bucket(extract_bucket_from_text(clause))
    return any(value != 0 for value in bucket.values())


def clause_mentions_attack_type(clause: str, row: AttackSummaryRow) -> bool:
    name_hits = [row.source_name, row.entity]
    if any(name and name in clause for name in name_hits):
        return True
    if row.row_type == "普攻" and "普攻" in clause:
        return True
    if row.row_type == "战技" and "战技" in clause:
        return True
    if row.row_type == "终结技" and "终结技" in clause:
        return True
    if row.row_type == "秘技" and "秘技" in clause:
        return True
    if "追加攻击" in row.tags and ("追加攻击" in clause or "神君" in clause or "乘胜追击" in clause):
        return True
    if "持续伤害" in row.tags and "持续伤害" in clause:
        return True
    if "召唤物" in row.tags and any(token in clause for token in ["忆灵", "召唤物", "死龙"]):
        return True
    if "欢愉伤害" in row.tags and "欢愉伤害" in clause:
        return True
    return False


def clause_is_generic(clause: str) -> bool:
    specific_tokens = [
        "普攻",
        "战技",
        "终结技",
        "追加攻击",
        "持续伤害",
        "欢愉伤害",
        "忆灵",
        "召唤物",
        "死龙",
        "神君",
        "乘胜追击",
        "燎尽黯泽的焰息",
        "擘裂冥茫的爪痕",
        "灼掠幽墟的晦翼",
    ]
    return not any(token in clause for token in specific_tokens)


def apply_effect_clauses(entries: Iterable[Entry], rows: List[AttackSummaryRow]) -> None:
    rows_by_role: Dict[str, List[AttackSummaryRow]] = defaultdict(list)
    for row in rows:
        rows_by_role[row.role].append(row)

    for role, role_rows in rows_by_role.items():
        if role in SUPPORT_ROLES:
            continue
        role_entries = [e for e in entries if e.role == role and e.source_col in ZERO_EIDOLON_SOURCES]
        for entry in role_entries:
            for clause in split_effect_clauses(entry.text):
                bucket = non_base_bucket(extract_bucket_from_text(clause, role=role, source_col=entry.source_col, source_name=entry.source_name))
                if not bucket:
                    continue
                for row in role_rows:
                    if row.source_name == entry.source_name and row.base_text == entry.text:
                        continue
                    applies = False
                    if role == "遐蝶" and "遐蝶与死龙" in clause:
                        applies = True
                    elif role == "遐蝶" and "死龙" in clause and "遐蝶" not in clause:
                        applies = "召唤物" in row.tags
                    elif role == "遐蝶" and "遐蝶" in clause and "死龙" not in clause:
                        applies = "召唤物" not in row.tags
                    elif clause_mentions_attack_type(clause, row):
                        applies = True
                    elif clause_is_generic(clause):
                        applies = True
                    if applies:
                        merge_bucket(row.effect_bucket, bucket)
                        row.notes.append(clause)


def build_attack_summary_rows(entries: Iterable[Entry]) -> List[AttackSummaryRow]:
    rows = build_base_rows(entries)
    rows.extend(build_support_rows(entries))
    rows.extend(fetch_castorice_servant_rows())
    apply_effect_clauses(entries, rows)
    order_map = {role: idx for idx, role in enumerate(TARGET_ROLES)}
    row_type_order = {"普攻": 1, "战技": 2, "终结技": 3, "追加攻击": 4, "持续伤害": 5, "秘技": 6, "召唤物": 7, "增益": 8}
    source_order = {value: idx for idx, value in enumerate(SOURCE_ORDER)}
    rows.sort(
        key=lambda row: (
            order_map.get(row.role, 999),
            row_type_order.get(row.row_type, 99),
            source_order.get(row.source_col, 999),
            row.source_name,
            row.base_text,
        )
    )
    return rows


def decimal_to_number(value: Decimal) -> int | float:
    if value == value.to_integral():
        return int(value)
    return float(value)


THIN_SIDE = Side(style="thin", color="D9DEE7")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="5B9BD5")
LABEL_FILL = PatternFill("solid", fgColor="EAF2F8")
ALT_FILL_1 = PatternFill("solid", fgColor="F8FBFF")
ALT_FILL_2 = PatternFill("solid", fgColor="F3F7F1")
NOTE_FILL = PatternFill("solid", fgColor="FFF4CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def apply_border_grid(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = TABLE_BORDER


def estimate_row_height(values: Iterable[object], chars_per_line: int = 22) -> float:
    line_count = 1
    for value in values:
        if value is None:
            continue
        logical_lines = 0
        for part in str(value).splitlines() or [""]:
            logical_lines += max(1, (len(part) + chars_per_line - 1) // chars_per_line)
        line_count = max(line_count, logical_lines)
    return min(max(22, 18 * line_count), 140)


def style_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    fill: PatternFill = HEADER_FILL,
) -> None:
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = CENTER_WRAP
        cell.border = TABLE_BORDER
    ws.row_dimensions[row_idx].height = 30


def apply_summary_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.freeze_panes = "I2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 80
    style_header_row(ws, 1)

    role_to_fill = {
        role: ALT_FILL_1 if idx % 2 == 0 else ALT_FILL_2
        for idx, role in enumerate(TARGET_ROLES)
    }
    for row_idx in range(2, ws.max_row + 1):
        fill = role_to_fill.get(ws.cell(row_idx, 1).value)
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 9)]
        ws.row_dimensions[row_idx].height = estimate_row_height(row_values, chars_per_line=30)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = TABLE_BORDER
            if col_idx <= 8 or col_idx == ws.max_column:
                cell.alignment = LEFT_WRAP
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 56
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 26
    for col_idx in range(9, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def apply_detail_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 85
    style_header_row(ws, 1, SECTION_FILL)

    preset_widths = {1: 12, 2: 10, 3: 22, 4: 20, 5: 64, 6: 40}
    for col_idx, width in preset_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(7, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 7)]
        ws.row_dimensions[row_idx].height = estimate_row_height(row_values, chars_per_line=28)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = TABLE_BORDER
            if col_idx <= 6 or col_idx == ws.max_column:
                cell.alignment = LEFT_WRAP
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_notes_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 88
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    style_header_row(ws, 1, SECTION_FILL)

    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, min(ws.max_column, 4) + 1)]
        ws.row_dimensions[row_idx].height = estimate_row_height(values, chars_per_line=36)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = TABLE_BORDER
            cell.alignment = LEFT_WRAP
            if row_idx == 8:
                cell.fill = LABEL_FILL
                cell.font = BOLD_FONT


def style_role_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "B13"

    max_col = ws.max_column
    max_row = ws.max_row

    stale_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= 2 <= merged_range.max_row or merged_range.min_row <= 1 <= merged_range.max_row:
            stale_ranges.append(str(merged_range))
    for merged_range in stale_ranges:
        ws.unmerge_cells(merged_range)

    if max_col >= 2:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    title_cell = ws.cell(1, 1)
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    desc_cell = ws.cell(2, 1)
    desc_cell.fill = NOTE_FILL
    desc_cell.alignment = LEFT_WRAP
    ws.row_dimensions[2].height = estimate_row_height([desc_cell.value], chars_per_line=80)

    for row_idx in range(4, min(10, max_row) + 1):
        ws.cell(row_idx, 1).fill = LABEL_FILL
        ws.cell(row_idx, 1).font = BOLD_FONT
        ws.cell(row_idx, 1).alignment = CENTER_WRAP
        ws.cell(row_idx, 2).alignment = LEFT_WRAP
    if max_row >= 4:
        apply_border_grid(ws, 4, min(10, max_row), 1, min(2, max_col))

    if max_row >= 12:
        for cell in ws[12]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = CENTER_WRAP
            cell.border = TABLE_BORDER
        ws.row_dimensions[12].height = 36

    upper_end = min(31, max_row)
    if upper_end >= 13:
        for row_idx in range(13, upper_end + 1):
            ws.cell(row_idx, 1).fill = LABEL_FILL
            ws.cell(row_idx, 1).font = BOLD_FONT
            ws.cell(row_idx, 1).alignment = CENTER_WRAP
            for col_idx in range(2, max_col + 1):
                ws.cell(row_idx, col_idx).alignment = LEFT_WRAP
            ws.row_dimensions[row_idx].height = estimate_row_height(
                [ws.cell(row_idx, col_idx).value for col_idx in range(2, max_col + 1)],
                chars_per_line=24,
            )
        apply_border_grid(ws, 12, upper_end, 1, max_col)

    if max_row >= 34:
        for cell in ws[34]:
            cell.fill = SECTION_FILL
            cell.font = WHITE_FONT
            cell.alignment = CENTER_WRAP
            cell.border = TABLE_BORDER
        ws.row_dimensions[34].height = 24

    if max_row >= 35:
        for row_idx in range(35, max_row + 1):
            ws.cell(row_idx, 1).fill = LABEL_FILL
            ws.cell(row_idx, 1).font = BOLD_FONT
            ws.cell(row_idx, 1).alignment = CENTER_WRAP
            for col_idx in range(2, max_col + 1):
                ws.cell(row_idx, col_idx).alignment = LEFT_WRAP
            ws.row_dimensions[row_idx].height = estimate_row_height(
                [ws.cell(row_idx, col_idx).value for col_idx in range(2, max_col + 1)],
                chars_per_line=26,
            )
        apply_border_grid(ws, 34, max_row, 1, max_col)


ROLE_SHEET_CATEGORY_TO_KEYS = {
    "基础倍率/附加伤害": BASE_KEYS,
    "攻击/生命/防御加成": {"攻击%", "生命%", "防御%", "固定攻击", "固定生命", "固定防御"},
    "速度": {"速度%", "速度点"},
    "增伤区": {"增伤区%"},
    "易伤区": {"易伤区%"},
    "暴击率": {"暴击率%"},
    "暴击伤害": {"暴击伤害%"},
    "防御穿透/减防": {"防御穿透/减防%"},
    "抗穿/减抗": {"抗穿/减抗%"},
    "击破特攻": {"击破特攻%"},
    "击破伤害提高": {"击破伤害提高%"},
    "超击破伤害提高": {"超击破伤害提高%"},
    "弱点击破效率/无视弱点削韧": {"弱点击破效率/无视弱点削韧%"},
    "欢愉度": {"欢愉度%"},
    "笑点": {"笑点"},
    "增笑": {"增笑%"},
    "伤害减免/受击减伤": {"伤害减免/受击减伤%"},
}


def apply_role_sheet_light_cone_override(ws: openpyxl.worksheet.worksheet.Worksheet, role: str) -> None:
    overrides = LIGHT_CONE_OVERRIDES.get(role)
    if not overrides:
        return

    light_cone_name = overrides[0][0]
    ws["B8"] = light_cone_name
    ws["A2"] = f"{role} 的专属光锥已改为角色原始专武。单元格保留原句，方便后续继续校对。"

    light_cone_col = None
    for col_idx in range(2, ws.max_column + 1):
        value = ws.cell(12, col_idx).value
        if isinstance(value, str) and value.startswith("专属光锥"):
            light_cone_col = col_idx
            break
    if light_cone_col is None:
        return

    ws.cell(12, light_cone_col).value = f"专属光锥\n{light_cone_name}"
    for row_idx in range(13, 32):
        ws.cell(row_idx, light_cone_col).value = None

    override_entries: List[Entry] = []
    for source_name, text, source_url in overrides:
        entry = Entry(role=role, path="", element="", source_col="专属光锥", source_name=source_name, text=text, source_url=source_url)
        extract_entry(entry)
        override_entries.append(entry)

    assigned_texts: set[str] = set()
    for row_idx in range(13, 31):
        label = ws.cell(row_idx, 1).value
        if label not in ROLE_SHEET_CATEGORY_TO_KEYS:
            continue
        matched = []
        for entry in override_entries:
            if any(key in entry.extracted for key in ROLE_SHEET_CATEGORY_TO_KEYS[label]):
                matched.append(entry.text)
                assigned_texts.add(entry.text)
        if matched:
            ws.cell(row_idx, light_cone_col).value = "\n".join(dict.fromkeys(matched))

    leftovers = [entry.text for entry in override_entries if entry.text not in assigned_texts]
    if leftovers:
        ws.cell(31, light_cone_col).value = "\n".join(dict.fromkeys(leftovers))

    for row_idx in range(26, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == "专属光锥":
            combined_text = " ".join(text for _, text, _ in overrides)
            ws.cell(row_idx, 2).value = combined_text
            ws.cell(row_idx, 3).value = overrides[0][2]
            break


def copy_role_sheet(src_wb: openpyxl.Workbook, dst_wb: Workbook, role: str) -> None:
    src_ws = src_wb[role]
    dst_ws = dst_wb.create_sheet(role)

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws[cell.coordinate]
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.border = copy(cell.border)
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.number_format = cell.number_format
                dst_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                dst_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                dst_cell.comment = copy(cell.comment)

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    style_role_sheet(dst_ws)
    apply_role_sheet_light_cone_override(dst_ws, role)


def write_summary_sheet(wb: Workbook, entries: Iterable[Entry]) -> None:
    ws = wb.active
    ws.title = "汇总"
    headers = ["角色", "条目类型", "来源列", "来源名称", "基础倍率说明", "适用增益说明", "作用对象", "持续回合"] + ATTACK_SUMMARY_EXPORT_COLUMNS + ["备注"]
    ws.append(headers)

    for item in build_attack_summary_rows(entries):
        merged_bucket = bucket_with_defaults()
        merge_bucket(merged_bucket, item.base_bucket)
        merge_bucket(merged_bucket, item.effect_bucket)
        effect_lines = []
        for key in ATTACK_SUMMARY_EXPORT_COLUMNS:
            effect_value = item.effect_bucket.get(key, Decimal("0"))
            if effect_value:
                effect_lines.append(f"{key}={decimal_to_number(effect_value)}")
        note_text = " | ".join(dict.fromkeys(item.notes))
        row = [
            item.role,
            item.row_type,
            item.source_col,
            item.source_name,
            item.base_text,
            "\n".join(effect_lines),
            item.target_scope,
            item.duration,
        ]
        row.extend(decimal_to_number(merged_bucket[column]) for column in ATTACK_SUMMARY_EXPORT_COLUMNS)
        row.append(note_text)
        ws.append(row)
    apply_summary_layout(ws)


def write_detail_sheet(wb: Workbook, entries: Iterable[Entry]) -> None:
    ws = wb.create_sheet("解析明细")
    headers = [
        "角色",
        "来源列",
        "来源名称",
        "原始分类",
        "文本",
        "来源URL",
        *SUMMARY_COLUMNS,
        "备注",
    ]
    ws.append(headers)

    for entry in entries:
        row = [
            entry.role,
            entry.source_col,
            entry.source_name,
            " / ".join(entry.original_categories),
            entry.text,
            entry.source_url,
        ]
        row.extend(decimal_to_number(entry.extracted.get(column, Decimal("0"))) for column in SUMMARY_COLUMNS)
        row.append(" | ".join(entry.notes))
        ws.append(row)
    apply_detail_layout(ws)


def write_notes_sheet(wb: Workbook, entries: Iterable[Entry]) -> None:
    ws = wb.create_sheet("说明")
    ws.append(["说明项", "内容"])
    ws.append(["口径", "汇总页只展示 0命；专属光锥计入。辅助角色另外按增益条目列出作用对象与持续回合。角色原始单表与解析明细继续保留，便于回查。"])
    ws.append(["汇总方式", "输出分为伤害条目与辅助增益条目；普攻、战技、终结技、追加攻击、持续伤害、召唤物分别列出，花火/阮•梅/布洛妮娅额外列出增益行。"])
    ws.append(["动态层数", "文本中明确写出“最多叠加X层/次”的，按上限折算，例如 20%*3=60%、30%*6=180%。"])
    ws.append(["专项补充", "遐蝶补入了死龙的忆灵技倍率：擘裂冥茫的爪痕、燎尽黯泽的焰息、灼掠幽墟的晦翼。"])
    ws.append(["专武修正", "希儿已改回【于夜色中】，布洛妮娅已改回【但战斗还未结束】；不再沿用 Prydwen 当期最优候选光锥。"])
    ws.append(["异常修正", "爻光 / 流萤 / 白厄 行迹表中的异常速度值已按常见行迹数值修正为 `速度+9` / `速度+5`。"])

    ws.append([])
    ws.append(["角色", "来源列", "来源名称", "备注"])
    for entry in entries:
        if not entry.notes:
            continue
        ws.append([entry.role, entry.source_col, entry.source_name, " | ".join(entry.notes)])
    apply_notes_layout(ws)


def autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_width: int = 48) -> None:
    for column in ws.columns:
        length = 0
        letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build_output(src_path: Path, output_path: Path) -> Path:
    src_wb = openpyxl.load_workbook(src_path, data_only=False)
    entries = load_unique_entries(src_path)

    out_wb = Workbook()
    write_summary_sheet(out_wb, entries)
    write_detail_sheet(out_wb, entries)
    write_notes_sheet(out_wb, entries)

    for role in TARGET_ROLES:
        copy_role_sheet(src_wb, out_wb, role)

    try:
        out_wb.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = output_path.with_name(f"{output_path.stem}_新版{output_path.suffix}")
        out_wb.save(fallback_path)
        return fallback_path


def main() -> None:
    cwd = Path.cwd()
    src_path = cwd / "角色乘区拆分.xlsx"
    output_path = cwd / "目标角色乘区汇总.xlsx"
    build_output(src_path, output_path)
    print(output_path)


if __name__ == "__main__":
    main()
